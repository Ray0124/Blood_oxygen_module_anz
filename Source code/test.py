from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
from scipy import signal
from matplotlib import cm


C1_PD = []
C2_PD = []
C3_PD = []
C4_PD = []
time  = []
wb = load_workbook('sample_format.xlsx')
sh = wb[wb.sheetnames[0]]

num = sh.max_row
        # print(sh.max_column)
        # 按行讀取所有資料，每一行的單元格放入一個元組中
rows = sh.rows
list_row=list(rows)
for row in list_row[1:]:  # 遍歷每列資料
    case = []  # 用於存放一行資料
    for c in row:  # 把每行的每個單元格的值取出來，存放到case裡
        case.append(c.value)
    time.append(case[0])
    C1_PD.append(case[1])
    C2_PD.append(case[2])
    C3_PD.append(case[3])
    C4_PD.append(case[4])

# print(C1_PD)
overlap_fac=0.9
fft_size=512
win_size=512
data=C1_PD
fs=1000
# overlap_fac = np.float32(1 - overlap_fac)
# hop_size = np.int32(np.floor(win_size * overlap_fac))
# total_segments = np.int32(np.ceil(len(C1_PD) / np.float32(hop_size)))
# window = np.hanning(win_size) * overlap_fac * 2
# inner_pad = np.zeros((fft_size * 2) - win_size)
# pad_end_size = fft_size
# proc = np.concatenate((data, np.zeros(pad_end_size)))
# result = np.empty((total_segments, fft_size), dtype=np.float32)
#
# X=np.arange(fft_size) / np.float32(fft_size * 2) * fs
# print('X:',end=' ')
# print(len(X))
#
# print('hop_size:',end=' ')
# print(hop_size)
# print('total_segments:',end=' ')
# print(total_segments)
# # print('window:',end=' ')
# # print(window)
# # print('inner_pad:',end=' ')
# # print(inner_pad)
# print('proc:',end=' ')
# print(proc)
# print('num:',end=' ')
# print(num)

freqs, times, Sxx = signal.spectrogram(np.array(data), fs=fs, window='hanning',nperseg=200, noverlap=0.5*200,detrend=False, scaling='spectrum')

# print(Sxx)
# print(Sxx.shape)
# print(freqs.shape)
# print(times.shape)
# plt.pcolormesh(times, freqs, Sxx, shading='auto')
# plt.colorbar()
# plt.ylabel('Frequency [Hz]')
# plt.xlabel('Time [s]')
# plt.show()

X=times
Y=freqs
Z=10*np.log10(Sxx/1e-06)
print('meshgrid前:')
print(X)
print(Y)
X,Y=np.meshgrid(X,Y)
print('meshgrid後:')
print(X)
print(Y)
print(Z)
fig = plt.figure()
ax = fig.gca(projection='3d')
ax.can_pan()
ax.can_zoom()
surf = ax.plot_surface(X, Y, Z, cmap=cm.plasma,linewidth=0, antialiased=False)
xmin, xmax = ax.get_xlim()
ymin, ymax = ax.get_ylim()

x_plane = np.linspace(0, 500, 50)
y_plane = np.linspace(0, 100, 50)
X_T, Y_T = np.meshgrid(x_plane, y_plane)
X_P, Y_P = np.meshgrid(X, Y)


print(np.shape(X))
print(np.shape(Y))
print(np.shape(Z))
plane= ax.plot(X[0],Y[0],Z[0],'b')
ax.set_zlim(0, 170)
ax.set_xlabel('time')
ax.set_ylabel('freq')
ax.set_zlabel('db')
fig.colorbar(surf, shrink=0.5, aspect=20,)
plt.show()

