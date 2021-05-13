import matplotlib
matplotlib.use("Qt5Agg")
from w import *
from Subwindow import *
from scipy import fft,signal,signal
import os
from openpyxl import load_workbook
from PyQt5.QtGui import QColor



color_map=['IR','Red','Green','Blue']



#
# def FFT_peak(x,y):
#     peaks, properties = find_peaks(y, height=20, width=1)
#     return x[peaks],y[peaks]

# def FFT_use(num,Data_s,sample_rate):
#     N = num
#     data_ac = Data_s - np.mean(Data_s)
#     fft_y = fft.fft(data_ac)  # 快速傅立葉變換
#     x = np.arange(0, sample_rate, sample_rate / N)  # 頻率個數
#     half_x = x[range(int(N / 2))]  # 取一半區間
#     abs_y = np.abs(fft_y)  # 取複數的絕對值，即複數的模(雙邊頻譜)
#     angle_y = np.angle(fft_y)  # 取複數的角度
#     normalization_y = abs_y / N  # 歸一化處理（雙邊頻譜）
#     normalization_half_y = normalization_y[range(int(N / 2))]  # 由於對稱性，只取一半區間（單邊頻譜）
#     return half_x,normalization_half_y


class MainWindow(QWidget):
    serial_data = pyqtSignal(object)
    serial_data_time = pyqtSignal(object, object)
    data_fft = pyqtSignal(object, object)
    data_wave = pyqtSignal(object,object, object,object)
    data_wave_db = pyqtSignal(object, object, object, object)
    data_stft = pyqtSignal(object, object)
    data_DC = pyqtSignal(object,object,object,object)
    data_PI = pyqtSignal(object,object,object,object,object)
    data_PI_avg = pyqtSignal(object)
    data_SPO2_use = pyqtSignal(object, object,object, object,object)
    data_HR = pyqtSignal(object,object, object, object)
    data_HR_value = pyqtSignal(object)
    data_welch = pyqtSignal(object,object,object,object)
    data_welch_value = pyqtSignal(object)
    data_power = pyqtSignal(object,object,object,object)
    data_power_value = pyqtSignal(object)
    def __init__(self,parent=None):
        super(self.__class__, self).__init__()

        self.ui=Ui_Form()
        self.ui.setupUi(self)
        self.main_widget = QtWidgets.QWidget(self)
        self.sub_window_intensity = SubWindow()
        self.sub_window_stft = SubWindow_STFT()
        self.sub_window_DC = SubWindow_DC()
        self.sub_window_PI = SubWindow_PI()
        self.sub_window_SPO2 = SubWindow_SPO2()
        self.sub_window_HR = SubWindow_HR()
        self.sub_window_power = SubWindow_power()
        self.ui.select_file_btn.clicked.connect(self.Load_file)
        self.ui.function_btn.clicked.connect(self.Start_s)
        self.ui.function_reset.clicked.connect(self.reset_f)
        self.serial_data.connect(self.sub_window_intensity.ui_sub.canvas.chart_update)
        self.serial_data_time.connect(self.sub_window_intensity.ui_sub.canvas.chart_update_time)
        self.data_fft.connect(self.sub_window_intensity.ui_sub.canvas.fft_chart)
        self.data_wave.connect(self.sub_window_intensity.ui_sub.canvas.plot_map)
        self.data_wave_db.connect(self.sub_window_intensity.ui_sub.canvas.plot_map_db)
        self.data_stft.connect(self.sub_window_stft.receive_data)
        self.data_DC.connect(self.sub_window_DC.receive_data)
        self.data_PI.connect(self.sub_window_PI.ui_sub_pi.canvas.chart_update_scatter)
        self.data_HR.connect(self.sub_window_HR.ui_sub_HR.canvas.HR_plot)
        self.data_SPO2_use.connect(self.sub_window_SPO2.receive_data)
        self.data_HR_value.connect(self.sub_window_HR.ui_sub_HR.lineEdit.setText)
        self.data_PI_avg.connect(self.sub_window_PI.ui_sub_pi.lineEdit.setText)
        self.data_welch.connect(self.sub_window_power.ui_sub_power.canvas.welch_power_plot)
        self.data_power.connect(self.sub_window_power.ui_sub_power.canvas.FFT_power_plot)
        self.data_welch_value.connect(self.sub_window_power.ui_sub_power.lineEdit.setText)
        self.data_power_value.connect(self.sub_window_power.ui_sub_power.lineEdit.setText)
        self.ui.function_btn.setEnabled(False)
        self.ui.comboBox_2.setEnabled(False)
        self.ui.lineEdit.setEnabled(False)
        self.ui.function_reset.setEnabled(False)
        # self.sub_window_PI.ui_sub_pi.lineEdit.setEnabled(False)
        # self.sub_window_power.ui_sub_power.lineEdit.setEnabled(False)
        # self.col = QColor(0, 0, 0)
        # self.square = QFrame(self)
        # self.square.setStyleSheet("QWidget { background-color: %s }" %self.col.name())


    def Load_file(self):
        self.cwd = os.getcwd()
        self.ui.comboBox.clear()
        self.Item = []
        self.time = []
        self.num=0
        self.test_data = []
        # self.sample_rate=0
        fileName_choose, filetype = QFileDialog.getOpenFileName(self,"選取文件",self.cwd,"All Files (*);;Text Files (*.txt)")
        if fileName_choose == "":
            print("\n取消")
            return
        # 開啟工作簿
        wb = load_workbook(fileName_choose)
        # 獲取表單
        sh = wb[wb.sheetnames[0]]
        self.num = sh.max_row
        rows = sh.rows
        list_row=list(rows)
        for c in list_row[0][1:]:
            self.Item.append(c.value)
        self.ui.comboBox.addItems(self.Item)
        lenc=len(self.Item)
        for i in range(lenc):
            self.test_data.append([])

        for row in list_row[1:]:  # 遍歷每列資料
            case = []  # 用於存放一行資料
            for c in row:  # 把每行的每個單元格的值取出來，存放到case裡
                case.append(c.value)
            self.time.append(case[0])
            for i in range(lenc):
                self.test_data[i].append(case[i + 1])


        self.ui.function_btn.setEnabled(True)
        self.ui.comboBox_2.setEnabled(True)
        self.ui.lineEdit.setEnabled(True)
        self.ui.function_reset.setEnabled(True)
        self.ui.select_file_btn.setEnabled(False)


    def Start_s(self):
        sample_rate = self.ui.lineEdit.text()
        if sample_rate == '':
            QMessageBox.warning(self, "錯誤", "SampleRate不可為空")
            return

        sample_rate = float(sample_rate)
        selection = self.ui.comboBox.currentText()
        selection_ways=self.ui.comboBox_2.currentText()
        target_index=self.Item.index(selection)
        # print(target_index)
        Data_s=self.test_data[target_index]
        # print(Data_s)
        data_ac = Data_s - np.mean(Data_s)




        if selection_ways == 'Intensity(points)':
            self.sub_window_intensity.show()
            self.serial_data.emit(Data_s)
        elif selection_ways == 'Intensity(time)':
            self.sub_window_intensity.show()
            self.serial_data_time.emit(self.time,Data_s)
        elif selection_ways == 'FFT':
            N = self.num
            fft_y = fft.fft(data_ac)  # 快速傅立葉變換
            x = np.arange(0,  sample_rate,  sample_rate / N)  # 頻率個數
            half_x = x[range(int(N / 2))]  # 取一半區間
            abs_y = np.abs(fft_y)  # 取複數的絕對值，即複數的模(雙邊頻譜)
            angle_y = np.angle(fft_y)  # 取複數的角度
            normalization_y = abs_y / N  # 歸一化處理（雙邊頻譜）
            normalization_half_y = normalization_y[range(int(N / 2))]  # 由於對稱性，只取一半區間（單邊頻譜）
            self.sub_window_intensity.show()
            self.data_fft.emit(half_x,normalization_half_y)

        elif selection_ways == 'SFFT(3d)':
            self.sub_window_stft.show()
            self.data_stft.emit(data_ac,sample_rate)


        elif selection_ways == 'Wavelet transform':
            self.sub_window_intensity.show()
            self.data_wave.emit(self.num,self.time, data_ac,sample_rate)

        elif selection_ways == 'Wavelet transform(db)':
            self.sub_window_intensity.show()
            self.data_wave_db.emit(self.num,self.time, data_ac,sample_rate)

        elif selection_ways == 'DC_anz':
            N = self.num
            self.sub_window_DC.show()
            self.data_DC.emit(Data_s, sample_rate,N,self.time)


        elif selection_ways == 'PI_anz':
            dis=self.ui.lineEdit_2.text()
            if dis=='':
                dis=50
            else:
                dis=int(dis)



            self.sub_window_PI.show()
            peaks, _ = signal.find_peaks(Data_s, distance=dis)
            re_Data_s = []
            for i in peaks:
                re_Data_s.append(Data_s[i])
            tmp_tem=[]
            for i in Data_s:
                tmp_tem.append(-i)
            valley_indexes,_=signal.find_peaks(tmp_tem, distance=50)
            re_Data_valley = []
            for i in valley_indexes:
                re_Data_valley.append(Data_s[i])


            self.data_PI.emit(peaks,re_Data_s,Data_s,valley_indexes,re_Data_valley)
            count_peak=peaks[1:len(peaks)-1]
            count_valley = valley_indexes[1:len(valley_indexes) - 1]

            PI=[]
            # for i in range(len(re_Data_s)):
            #     print(re_Data_s[i])
            #     print(re_Data_valley[i])

            if len(count_peak)>len(count_valley):
                for i in range(len(count_valley)):
                    PI.append((re_Data_s[i]-re_Data_valley[i])/re_Data_valley[i])
            else:
                for i in range(len(count_peak)):
                    PI.append((re_Data_s[i]-re_Data_valley[i])/re_Data_valley[i])

            PI_avg=np.mean(PI)
            PI_avg=round(PI_avg, 3)*100
            PI_avg = round(PI_avg, 3)
            self.data_PI_avg.emit(str(PI_avg))


        elif selection_ways == 'SPO2':
            self.sub_window_SPO2.show()
            self.data_SPO2_use.emit(self.test_data[0],self.test_data[1],sample_rate,self.num,self.time)


        elif selection_ways == 'Heart_Rate':
            self.sub_window_HR.show()

            dis = self.ui.lineEdit_2.text()
            if dis == '':
                dis = 50
            else:
                dis = int(dis)
            peaks, _ = signal.find_peaks(Data_s, distance=dis)
            re_Data_s = []
            for i in peaks:
                re_Data_s.append(Data_s[i])
            peak_time=[]
            T=float(1/sample_rate)
            for i in peaks:
                peak_time.append(i*T)

            self.data_HR.emit(self.time,Data_s,peak_time,re_Data_s)

            HR_v=round(len(re_Data_s)/(T*self.num)*60,3)
            self.data_HR_value.emit(str(HR_v))


        elif selection_ways == 'Spec_Power(welch)':
            self.sub_window_power.show()
            f, Pxx_spec = signal.welch(Data_s, sample_rate, 'flattop', 256, scaling='spectrum')
            self.data_welch.emit(f, Pxx_spec,0.8,2.5)

            f_index=[]
            f_index_value = []
            sum=0
            for i in range(len(f)):
                sum=sum+Pxx_spec[i]

            range_sum=0
            for i in range(len(f)):
                if 0.8<=f[i]<=2.5:
                    range_sum=range_sum+Pxx_spec[i]

            SNR=round(10*np.log10(float(range_sum/(sum-range_sum))),3)
            self.data_welch_value.emit(str(SNR))




        elif selection_ways == 'Spec_Power(FFT)':
            self.sub_window_power.show()

            N = self.num
            fft_y = fft.fft(data_ac)  # 快速傅立葉變換
            x = np.arange(0, sample_rate, sample_rate / N)  # 頻率個數
            half_x = x[range(int(N / 2))]  # 取一半區間
            abs_y = np.abs(fft_y)  # 取複數的絕對值，即複數的模(雙邊頻譜)
            angle_y = np.angle(fft_y)  # 取複數的角度
            normalization_y = abs_y/N  # 歸一化處理（雙邊頻譜）
            normalization_half_y = normalization_y[range(int(N / 2))]  # 由於對稱性，只取一半區間（單邊頻譜）
            normalization_half_y_power=[]
            for i in normalization_half_y:
                normalization_half_y_power.append(i*i)

            self.data_power.emit(half_x, normalization_half_y_power, 0.8, 2.5)
            sum = 0
            for i in range(len(half_x)):
                sum = sum + normalization_half_y_power[i]

            range_sum = 0
            for i in range(len(half_x)):
                if 0.8 <= half_x[i] <= 2.5:
                    range_sum = range_sum + normalization_half_y_power[i]

            SNR = round(10 * np.log10(float(range_sum / (sum - range_sum))), 3)



            self.data_power_value.emit(str(SNR))



    def reset_f(self):
        self.ui.select_file_btn.setEnabled(True)
        self.ui.function_reset.setEnabled(False)
        self.ui.function_btn.setEnabled(False)


