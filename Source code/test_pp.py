from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
from scipy import signal
from matplotlib import cm
import scipy



def FFT_peak(x,y):
    peaks, properties = signal.find_peaks(y, height=20, width=1)
    return x[peaks],y[peaks]

C1_PD = []
C2_PD = []
C3_PD = []
C4_PD = []
time  = []
wb = load_workbook('sample_format.xlsx')
sh = wb[wb.sheetnames[0]]

num = sh.max_row
        # print(sh.max_column)
        # 按行讀取所有資料，每一行的單元格放入一個元組中
rows = sh.rows
list_row=list(rows)
for row in list_row[1:]:  # 遍歷每列資料
    case = []  # 用於存放一行資料
    for c in row:  # 把每行的每個單元格的值取出來，存放到case裡
        case.append(c.value)
    time.append(case[0])
    C1_PD.append(case[1])
    C2_PD.append(case[2])
    C3_PD.append(case[3])
    C4_PD.append(case[4])


plt.plot(time,C2_PD)
x,y=FFT_peak(time,C2_PD)
plt.plot(x,y)
plt.plot(np.zeros_like(C2_PD), "--", color="gray")
plt.show()

plt.show()


